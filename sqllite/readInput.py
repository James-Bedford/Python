#!/usr/bin/python

#Load the openpyxl library
from openpyxl import load_workbook
from optparse import OptionParser
import os

parser = OptionParser()

#Set some flags for all required parameters
parser.add_option("-a", "--append", action="store_true", help = "Append contents of CSV file to an existing database rather than creating a new one.")
parser.add_option("-r", "--disable-apparmor", action="store_true", help = "Disable apparmor.")
parser.add_option("-c", "--csvDir", type="string", help = "Specifiy the directory in which CSV files are to be stored.")
parser.add_option("-d", "--db", type="string", help = "Which database will be imported to. If 'none', disable SQL.")
parser.add_option("-s", "--schema", type="string", help = "Schema file to use for creating database.")
parser.add_option("-u", "--dbUser", type="string", help = "The MySQL user that will access the database.")
parser.add_option("-x", "--xlsxFile", type="string", help = "The xlsx file that contains all required camera information.")

(options, args) = parser.parse_args()

#Check that each option has been passed in and exit if not
if not options.csvDir:
    print "[ERROR] CSV file has not been specified and must be passed in via -c before running"
    quit()
else:
    csvDir = options.csvDir

if not options.db:
    print "[ERROR] Database has not been specified and must be passed in via -d before running"
    quit()
else:
    db = options.db

if not options.schema and options.db != "none":
    print "[ERROR] Database schema file not been specified and must be passed in via -s before running"
    quit()
else:
    dbSchema = options.schema

if not options.dbUser and options.db != "none":
    print "[ERROR] Database usr has not been specified and must be passed in via -u before running"
    quit()
else:
    dbUser = options.dbUser

if not options.xlsxFile:
    print "[ERROR] XLSX file has not been specified and must be passed in via -x before running"
    quit()
else:
    xlsxFile = options.xlsxFile

#Load the camera spreadsheet as a workbook
wb = load_workbook(filename = csvDir + xlsxFile, use_iterators = True)

#Get all the sheets in xlsx file and use them. Make sure each sheet has a corresponding entry in importData.sh as it contains CSV import statements which must be accurate
sheets = []
for sheet in wb.get_sheet_names():
    sheets.append(sheet)

#---------------------------------------------------------------
def format(csvFile):
    #Initialise wholeRow variable which will be used to format CSV strings
    wholeRow = ""

    #Iterate through every row and column excluding the first which is used for headers and ending at the maximum row value
    for row in ws.iter_rows("A2:"+str(chr(ws.max_column + 64))+str(ws.max_row)):
        for cell in row:

            #Create a CSV string usig the cell value and prefixing double quotes whilst postfixing double quotes and a comma
            if cell.value is None:
                wholeRow = wholeRow + "\"\","
            else:
                wholeRow = wholeRow + "\"" + str(cell.value) + "\","

        #When the CSV string has been finished remove the final comma
        wholeRow = wholeRow[:-1]+"\n"
        f = open(csvFile,"a")
        f.write(wholeRow)

        #Reset wholeRow variable after cells in row have been read
        wholeRow = ""
    try:
        f.close()
    except UnboundLocalError:
        f = open(csvFile,"a")
        f.write("")
        f.close()

#---------------------------------------------------------------
def csvWrite(sheet):
    columns = ""
    csvFile = csvDir+sheet+".csv"
    ws = wb.get_sheet_by_name(name = sheet)

    #Iterate through all cells but only in the first row of the provided xlsx file in order to get the column names
    for row in ws.iter_rows("A1:"+str(chr(ws.max_column + 64))+"1"):
        for cell in row:

            #Prepend command information to string if string is empty
            if columns == "":
                columns = "load data infile '"+csvDir+sheet+".csv' into table "+sheet+" fields terminated by ',' enclosed by '\"' ("+str(cell.value)

            #Otherwise add the extra field
            else:
                columns = columns+", "+str(cell.value)

        #Once finished append a closing parenthesis to end command
        columns = columns +")"
    global csvString
    csvString = columns

#---------------------------------------------------------------
def createDb(dbUser):
    import MySQLdb

    #Establish a MySQL connection without specifying a database and create the user defined database
    conn = MySQLdb.connect(user=dbUser)
    cur = conn.cursor()
    cur.execute("create database "+db)

    #Import master database schema to newly created database
    os.system("mysql -u "+dbUser+" "+db+" < "+dbSchema)
    conn.commit()
    conn.close()

#---------------------------------------------------------------
def apparmor(status):
    if not options.disable_apparmor:
        #Apparmor needs to be turned off if reading and writing from local files so sturn it off at the start and then turn it back on when finished
        if status == "stop":
            os.system("sudo service apparmor stop")
            os.system("sudo service apparmor teardown")
        elif status == "start":
            os.system("sudo service apparmor start")

#---------------------------------------------------------------
def checkDb(dbUser, db, state):
    import MySQLdb

    if state == "new":
        try:
            conn = MySQLdb.connect(user=dbUser, db=db)
            print "\n[WARNING] "+db+" database already exists.\n"
            cur = conn.cursor()
            conn.close()
            apparmor("start")
            quit()
        except MySQLdb.Error:
            #Create the database if it doesn't already exist
            createDb(dbUser)

    elif state == "append":
        try:
            conn = MySQLdb.connect(user=dbUser, db=db)
            cur = conn.cursor()
            conn.close()
        except MySQLdb.Error:
            print "\n[WARNING] "+db+" database does not exist.\n"
            apparmor("start")
            quit()

#---------------------------------------------------------------
#Main
#Stop apparmor
apparmor("stop")

#Create a new MySQL connection pointing to created database
if db != "none":
    import MySQLdb

    #Check to see if the database already exists
    if options.append:
        checkDb(dbUser, db, "append")
    else:
        checkDb(dbUser, db, "new")

    conn = MySQLdb.connect(user=dbUser, db=db)
    cur = conn.cursor()
else:
    conn = None
    cur = None

#Iterate through all sheets in the provided xlsx file
for sheet in sheets:
    csvString = ""
    csvWrite(sheet)
    #Csv file is just the dir and sheet name appended with .csv
    csvFile = csvDir+sheet+".csv"

    #Remove existing copy of csv file if it already exists
    if str(os.path.exists(csvFile)) == "True":
        os.remove(csvFile)

    #Get the specific sheet from withn xlsx file
    ws = wb.get_sheet_by_name(name = sheet)

    #Call format function for each sheet
    format(csvFile)

    #Call the previously formatted MySQL query to import csv file before removing that file
    if cur:
        cur.execute(csvString)
        os.remove(csvDir+sheet+".csv")

#Finally create the cameras table using prevoiusly entered information
if cur:
    cur.execute("delete from cameras")
    cur.execute("select cameraID, cameraName from cameraDetails into outfile '/tmp/table1.csv' fields terminated by ',' enclosed by '\"' lines terminated by '\n'")
    cur.execute("load data infile '/tmp/table1.csv' into table cameras fields terminated by ',' enclosed by '\"' (cameraID, cameraName);")
    os.system("sudo rm -f /tmp/table1.csv")

#Resume apparmor
apparmor("start")

#Close the database connection
if conn:
    conn.commit()
    conn.close()
